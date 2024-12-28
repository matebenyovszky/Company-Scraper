import pandas as pd
import os
import healing_agent
from search_engines import Google, Bing, Yahoo, Aol, Duckduckgo, Startpage
import openpyxl
from playwright.sync_api import sync_playwright
import re
import time
from datetime import datetime
import zipfile
from loguru import logger
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from threading import Thread, Event
from PIL import Image, ImageTk

logger.remove()
logger.add("scraper.log", format="{time:YYYY-MM-DD HH:mm:ss} | {level} | {message}", level="INFO")

browser = None
context = None
playwright = None


class ScraperApp:

    def __init__(self):
        self.root = tk.Tk()
        self.root.title('Website Scraper')
        self.root.geometry('600x600')
        self.file_frame = ttk.Frame(self.root, padding='10')
        self.file_frame.pack(fill=tk.X)
        self.file_label = ttk.Label(self.file_frame, text='Excel file:')
        self.file_label.pack(side=tk.LEFT)
        self.file_path = tk.StringVar()
        self.file_entry = ttk.Entry(self.file_frame, textvariable=self.
            file_path)
        self.file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.browse_btn = ttk.Button(self.file_frame, text='Browse',
            command=self.browse_file)
        self.browse_btn.pack(side=tk.LEFT)
        self.info_frame = ttk.Frame(self.root, padding='10')
        self.info_frame.pack(fill=tk.X)
        self.row_label = ttk.Label(self.info_frame, text='Current Row: -')
        self.row_label.pack()
        self.company_label = ttk.Label(self.info_frame, text='Company: -')
        self.company_label.pack()
        self.url_label = ttk.Label(self.info_frame, text='URL: -')
        self.url_label.pack()
        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(self.root, variable=self.
            progress_var, maximum=100)
        self.progress.pack(fill=tk.X, padx=10, pady=10)
        self.btn_frame = ttk.Frame(self.root, padding='10')
        self.btn_frame.pack()
        self.start_btn = ttk.Button(self.btn_frame, text='Start Processing',
            command=self.start_processing)
        self.start_btn.pack(side=tk.LEFT, padx=5)
        self.stop_btn = ttk.Button(self.btn_frame, text='Stop', command=
            self.stop_processing, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)
        self.timing_frame = ttk.Frame(self.root, padding='10')
        self.timing_frame.pack(fill=tk.X)
        self.timing_label = ttk.Label(self.timing_frame, text=
            'Last 5 website search times: -')
        self.timing_label.pack()
        self.screenshot_timing_label = ttk.Label(self.timing_frame, text=
            'Last 5 screenshot taking times: -')
        self.screenshot_timing_label.pack()
        self.last_found_frame = ttk.Frame(self.root, padding='10')
        self.last_found_frame.pack(fill=tk.X)
        self.last_found_label = ttk.Label(self.last_found_frame, text=
            'Last Found:')
        self.last_found_label.pack()
        self.last_found_info = ttk.Label(self.last_found_frame, text=
            'Row: -, Company: -, URL: -')
        self.last_found_info.pack()
        self.screenshot_label = ttk.Label(self.last_found_frame)
        self.screenshot_label.pack()
        self.processing = False
        self.stop_event = Event()
        self.last_5_get_official_url_times = []
        self.last_5_get_screenshot_url_times = []

    def browse_file(self):
        filename = filedialog.askopenfilename(filetypes=[('Excel files',
            '*.xlsx *.xls')])
        if filename:
            self.file_path.set(filename)

    def update_progress(self, current, total, company_name, url=None):
        if self.stop_event.is_set():
            return
        progress = current / total * 100
        self.progress_var.set(progress)
        self.row_label.config(text=f'Current Row: {current} of {total}')
        self.company_label.config(text=f'Company: {company_name}')
        self.url_label.config(text=f"URL: {url if url else '-'}")
        self.root.update()

    def update_timing_info(self):
        self.timing_label.config(text=
            f'Last 5 website search times: {self.last_5_get_official_url_times}'
            )
        self.screenshot_timing_label.config(text=
            f'Last 5 screenshot taking times: {self.last_5_get_screenshot_url_times}'
            )
        self.root.update()

    def update_last_found_info(self, row, company_name, url, screenshot_path):
        self.last_found_info.config(text=
            f'Row: {row}, Company: {company_name}, URL: {url}')
        if screenshot_path and os.path.exists(screenshot_path):
            image = Image.open(screenshot_path)
            image = image.resize((300, 300), Image.LANCZOS)
            photo = ImageTk.PhotoImage(image)
            self.screenshot_label.config(image=photo)
            self.screenshot_label.image = photo
        else:
            self.screenshot_label.config(image='')
            self.screenshot_label.image = None
        self.root.update()

    def start_processing(self):
        if not self.file_path.get():
            messagebox.showerror('Error', 'Please select an input file')
            return
        self.processing = True
        self.stop_event.clear()
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        Thread(target=self.process_file, daemon=True).start()

    def stop_processing(self):
        self.stop_event.set()
        self.processing = False
        cleanup_browser()
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        messagebox.showinfo('Stopped', 'Processing has been stopped')

    def process_file(self):
        try:
            df = process_companies_file(self.file_path.get(),
                progress_callback=self.update_progress, stop_event=self.
                stop_event, timing_callback=self.update_timing_info,
                last_5_get_official_url_times=self.
                last_5_get_official_url_times,
                last_5_get_screenshot_url_times=self.
                last_5_get_screenshot_url_times, last_found_callback=self.
                update_last_found_info)
            if not self.stop_event.is_set():
                messagebox.showinfo('Success', 'Processing completed!')
        except Exception as e:
            messagebox.showerror('Error', str(e))
        finally:
            self.stop_processing()

    def run(self):
        self.root.mainloop()


def ensure_browser():
    global browser, context, playwright
    if not browser:
        playwright = sync_playwright().start()
        browser = playwright.firefox.launch()
        context = browser.new_context(viewport={'width': 1280, 'height': 
            1280}, user_agent=
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36'
            , device_scale_factor=1, has_touch=False, is_mobile=False,
            java_script_enabled=True)


def cleanup_browser():
    global browser, context, playwright
    if browser:
        browser.close()
        browser = None
        context = None
    if playwright:
        playwright.stop()
        playwright = None


@healing_agent
def get_official_url(company_name, city, search_engine_starters=[Google(),
    Bing(), Yahoo(), Aol(), Duckduckgo(), Startpage()]):
    """
    Get the official website URL for a company by searching for it online.
    Uses multiple search engines and filters out blacklisted domains.
    
    Args:
        company_name: Name of the company
        city: City where company is located
        
    Returns:
        str: Official website URL if found, "na" if blocked, None otherwise
    """
    # these will be an "na" result
    blacklist = ['opten.hu', 'linkedin.com', 'nemzeticegtar.hu', 'emis.com', 'cinando.com',
        'instagram.com', 'ceginformacio.hu', 'ceginfo.hu', 'twitter.com',
        'youtube.com', 'pinterest.com', 'glassdoor.com', 'indeed.com',
        'bloomberg.com', 'crunchbase.com', 'yelp.com', 'yellowpages.com',
        'bbb.org', 'manta.com', 'bizapedia.com', 'google.hu', 'google.com',
        'cylex.hu', 'nemzeticegtar.hu', 'onlinecegnyilvantarto.hu',
        'wikipedia.org', 'companywall.hu', 'emis.com', 'economx.hu',
        'portfolio.hu', 'mfor.hu', 'kozlonyok.hu', 'trademagazin.hu',
        'magyarcegnyilvantarto.hu', 'topcegkereso.hu', 'xir.hu', 'facebook.com',
        'cegcontrol.hu', 'cegtalalo.hu', 'k-monitor.hu', 'dnb.com','www.sztnh.gov.hu', 'www.vasmegye.hu', 'nagykovacsi.hu']
    
    # these are deleted from the filtered list, so not considered as a result
    blocklist = ['google.hu', 'google.com', 'szotar.net', 'linkedin.com', 'fliphtml5.com', 'www.ceginformacio.hu', 'facebook.com']
    # Remove special characters from company name that could interfere with search
    company_name = company_name.replace('"', '')
    company_name = company_name.replace("'", '')
    company_name = company_name.replace('„', '')
    company_name = company_name.replace('(', '')
    company_name = company_name.replace(')', '')
    query = f'{company_name} hivatalos weblap'
    logger.info(f'Searching for {query}')

    # Try ScraperAPI first
    try:
        import requests
        payload = {
            'api_key': 'xxx', #your key here
            'query': query,
            'num': '10'
        }
        r = requests.get('https://api.scraperapi.com/structured/google/search', params=payload)
        data = r.json()
        
        if 'organic_results' in data:
            urls = [result['link'] for result in data['organic_results']]
            if urls:
                filtered_urls = [url for url in urls if not any(blocked in url.lower() for blocked in blocklist)]
                logger.info(f'Filtered URLs from ScraperAPI: {filtered_urls}')
                if filtered_urls:
                    first_url = filtered_urls[0]
                    domain = first_url.split('/')[2].lower()
                    if any(domain == blocked for blocked in blacklist):
                        logger.info(f'First ScraperAPI result for {company_name} was blocked: {first_url}')
                        return 'na'
                    parts = first_url.split('/')
                    if len(parts) > 3 and 'hu' in parts[3]:
                        main_url = '/'.join(parts[:4])
                    else:
                        main_url = '/'.join(parts[:3])
                    logger.info(f'Found URL via ScraperAPI for {company_name}: {main_url}')
                    return main_url
    except Exception as e:
        logger.error(f'Error with ScraperAPI: {str(e)}')

    # Fall back to other search engines if ScraperAPI fails
    for engine in search_engine_starters:
        try:
            encoded_query = query.encode('utf-8').decode('utf-8')
            engine.ignore_duplicate_urls = True
            engine.ignore_duplicate_domains = True
            results = engine.search(encoded_query, pages=1)
            urls = results.links()
            if not urls:
                return 'na'
            filtered_urls = [url for url in urls if not any(blocked in url.lower() for blocked in blocklist)]
            logger.info(f'Filtered URLs: {filtered_urls}')
            if not filtered_urls:
                return 'na'
            urls = filtered_urls
            first_url = urls[0]
            domain = first_url.split('/')[2].lower()
            if any(domain == blocked for blocked in blacklist):
                logger.info(
                    f'First result for {company_name} was blocked: {first_url}'
                    )
                return 'na'
            parts = first_url.split('/')
            if len(parts) > 3 and 'hu' in parts[3]:
                main_url = '/'.join(parts[:4])
            else:
                main_url = '/'.join(parts[:3])
            logger.info(f'Found URL for {company_name}: {main_url}')
            return main_url
        except UnicodeEncodeError as ue:
            logger.warning(
                f'Unicode encoding error occurred with search engine {engine.__class__.__name__}. '
                f'Unable to encode query "{query}" - error details: {str(ue)}. '
                f'This can happen with special characters in company names. '
                f'Skipping this search engine and trying next one.'
            )
            continue
        except Exception as e:
            logger.error(f'Error with {engine.__class__.__name__}: {str(e)}')
            continue
    return None


@healing_agent
def get_screenshot_url(official_url, row_number):
    """
    Navigate to URL, handle cookie popups, take screenshot and save it using Playwright.
    Includes anti-blocking measures and detection.
    
    Args:
        official_url: Website URL to screenshot
        row_number: Excel row number for filename
        
    Returns:
        str: Screenshot filename if successful, None otherwise
    """
    global context
    start_time = time.time()
    if not official_url:
        return None
    screenshots_dir = 'screenshots'
    if not os.path.exists(screenshots_dir):
        os.makedirs(screenshots_dir)
    domain = official_url.split('//')[-1].split('/')[0]
    clean_domain = re.sub('[^\\w\\-_]', '', domain)
    filename = f'{row_number}_{clean_domain}.jpg'
    filepath = os.path.join(screenshots_dir, filename)
    page = None
    try:
        logger.info(
            f'Starting browser setup at {time.time() - start_time:.2f}s')
        ensure_browser()
        page = context.new_page()
        logger.info(
            f'Browser setup completed at {time.time() - start_time:.2f}s')

        def handle_route(route):
            headers = route.request.headers
            headers['Accept'] = (
                'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'
                )
            headers['Accept-Language'] = 'hu-HU,hu;q=0.9,en;q=0.8'
            route.continue_(headers=headers)
        page.route('**/*', handle_route)
        logger.info(f'Starting page load at {time.time() - start_time:.2f}s')
        try:
            page.goto(official_url, wait_until='networkidle', timeout=25000)
            logger.info(f'Page loaded at {time.time() - start_time:.2f}s')
        except Exception as e:
            logger.error(f'Page load timeout for {official_url}: {str(e)}')
            logger.info('Taking screenshot of current page state')
        cookie_selectors = ["button:has-text('Accept')",
            "button:has-text('Elfogad')", "button:has-text('Elfogadom')",
            "button:has-text('Rendben')", "button:has-text('Megértettem')",
            "button:has-text('Hozzájárulok')", "button:has-text('Bezár')",
            "button:has-text('Accept Cookies')",
            "button:has-text('I Accept')", "button:has-text('Allow All')",
            "button:has-text('Allow Cookies')",
            "button:has-text('ÖSSZES')", "button:has-text('Összes')",
            "button:has-text('Összes süti elfogadása')",
            "button:has-text('OK')",
            "button:has-text('Elfogadása')",
            "button:has-text('Mindent elfogadok')",
            "button:has-text('Alle Akzeptieren')",
            "button:has-text('ELFOGADOM A JAVASOLT BEÁLLÍTÁSOKAT')",
            "button:has-text('AGREE AND CLOSE')",
            "button:has-text('Beállítások mentése')",
            "button:has-text('Az összes süti elfogadása')",
            "button:has-text('Javasolt beállítások elfogadása')",
            "button:has-text('Confirm')", '[id*=cookie] button',
            '[class*=cookie] button', '[id*=cookie-accept]',
            '[class*=cookie-accept]', '[id*=cookie-consent] button',
            '[class*=cookie-consent] button', '[id*=cookie-banner] button',
            '[id*=cookie-policy] button', '[class*=cookie-policy] button',
            '[id*=gdpr] button', '[class*=gdpr] button',
            '[id*=privacy] button', '[class*=privacy] button',
            "[aria-label*='cookie'] button", '[data-cookiebanner] button',
            '#onetrust-accept-btn-handler', 'button#accept-all-cookies',
            "button:has-text('Összes süti elfogadása')",
            'a#hs-eu-confirmation-button',
            'a#CybotCookiebotDialogBodyLevelButtonLevelOptinAllowAll',
            'button#c-p-bn',
            'button#CybotCookiebotDialogBodyLevelButtonLevelOptinAllowAll',
            'a.accept[onclick*="CloseOkEuCookieInfo"]',
            'a.cmpboxbtn.cmpboxbtnyes',
            'div.grey-box-btn.accept-default']
        logger.info(
            f'Starting cookie handling at {time.time() - start_time:.2f}s')
        for selector in cookie_selectors:
            try:
                if page.is_visible(selector):
                    page.click(selector, timeout=250)
                    page.wait_for_selector(selector, state='hidden',
                        timeout=5000)
                    break
            except Exception as e:
                logger.warning(
                    f'Error handling cookie selector {selector}: {str(e)}')
                continue
        logger.info(
            f'Cookie handling completed at {time.time() - start_time:.2f}s')
        logger.info(f'Taking screenshot at {time.time() - start_time:.2f}s')
        page.screenshot(path=filepath)
        logger.info(
            f'Screenshot saved to {filepath} at {time.time() - start_time:.2f}s'
            )
        return filename
    except Exception as e:
        logger.error(f'Screenshot error for {official_url}: {str(e)}')
        return None
    finally:
        if page:
            page.close()


@healing_agent
def process_companies_file(input_file, update_urls=True, update_screenshots
    =True, progress_callback=None, stop_event=None, timing_callback=None,
    last_5_get_official_url_times=None, last_5_get_screenshot_url_times=
    None, last_found_callback=None):
    """
    Process Excel file containing company information and add official_url
    and screenshot_url columns where missing. Uses openpyxl for better performance.
    
    Args:
        input_file: Path to Excel file
        update_urls: Whether to update missing official URLs
        update_screenshots: Whether to update missing screenshot URLs
        progress_callback: Callback function to update progress UI
        stop_event: Threading Event to signal stopping
        timing_callback: Callback function to update timing info UI
        last_5_get_official_url_times: List to store last 5 get_official_url times
        last_5_get_screenshot_url_times: List to store last 5 get_screenshot_url times
        last_found_callback: Callback function to update last found info UI
    """
    if not os.path.exists(input_file):
        raise FileNotFoundError(f'Input file not found: {input_file}')
    workbook = None
    try:
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        official_url_idx = headers.index('official_url'
            ) if 'official_url' in headers else None
        screenshot_url_idx = headers.index('screenshot_url'
            ) if 'screenshot_url' in headers else None
        company_name_idx = headers.index('cégnév')
        city_idx = headers.index('város')
        if official_url_idx is None:
            official_url_idx = len(headers)
            headers.append('official_url')
            sheet.cell(row=1, column=official_url_idx + 1, value='official_url'
                )
        if screenshot_url_idx is None:
            screenshot_url_idx = len(headers)
            headers.append('screenshot_url')
            sheet.cell(row=1, column=screenshot_url_idx + 1, value=
                'screenshot_url')
        total_rows = sheet.max_row
        logger.info(f'Total rows: {total_rows}')
        chunk_size = 5
        for row_idx in range(2, total_rows + 1, chunk_size):
            if stop_event and stop_event.is_set():
                logger.info('Processing stopped by user')
                break
                
            # Restart browser session every 20 rows
            if (row_idx - 2) % 20 == 0:
                logger.info('Restarting browser session after 20 rows')
                cleanup_browser()
                ensure_browser()
                
            chunk_end = min(row_idx + chunk_size, total_rows + 1)
            updates_made = False
            for current_row in range(row_idx, chunk_end):
                if stop_event and stop_event.is_set():
                    break
                logger.info('')
                logger.info(f'🏟️  Processing row {current_row} of {total_rows}'
                    )
                company_name = sheet.cell(row=current_row, column=
                    company_name_idx + 1).value
                city = sheet.cell(row=current_row, column=city_idx + 1).value
                official_url = sheet.cell(row=current_row, column=
                    official_url_idx + 1).value
                screenshot_url = sheet.cell(row=current_row, column=
                    screenshot_url_idx + 1).value
                if progress_callback:
                    progress_callback(current_row - 1, total_rows - 1,
                        company_name, official_url)
                if update_urls and not official_url:
                    start_time = time.time()
                    logger.info(f'Getting official URL for {company_name}...')
                    new_url = get_official_url(company_name, city)
                    elapsed = round(time.time() - start_time, 1)
                    logger.info(f'Getting official URL took {elapsed} seconds')
                    if last_5_get_official_url_times is not None:
                        last_5_get_official_url_times.append(elapsed)
                        if len(last_5_get_official_url_times) > 5:
                            last_5_get_official_url_times.pop(0)
                        if timing_callback:
                            timing_callback()
                    if new_url:
                        sheet.cell(row=current_row, column=official_url_idx +
                            1, value=new_url)
                        updates_made = True
                        official_url = new_url
                        if progress_callback:
                            progress_callback(current_row - 1, total_rows -
                                1, company_name, official_url)
                if (update_screenshots and not screenshot_url and
                    official_url and official_url != 'na'):
                    start_time = time.time()
                    logger.info(f'Getting screenshot for {official_url}...')
                    new_screenshot = get_screenshot_url(official_url,
                        current_row)
                    elapsed = round(time.time() - start_time, 1)
                    logger.info(f'Getting screenshot took {elapsed} seconds')
                    if last_5_get_screenshot_url_times is not None:
                        last_5_get_screenshot_url_times.append(elapsed)
                        if len(last_5_get_screenshot_url_times) > 5:
                            last_5_get_screenshot_url_times.pop(0)
                        if timing_callback:
                            timing_callback()
                    if new_screenshot:
                        sheet.cell(row=current_row, column=
                            screenshot_url_idx + 1, value=new_screenshot)
                        updates_made = True
                        if last_found_callback:
                            last_found_callback(current_row, company_name,
                                official_url, os.path.join('screenshots',
                                new_screenshot))
                logger.info(f'Processed row {current_row} of {total_rows}')
            if updates_made:
                workbook.save(input_file)
        return pd.read_excel(input_file)
    except FileNotFoundError as fnf_error:
        raise FileNotFoundError(f'Input file issue: {str(fnf_error)}')
    except openpyxl.utils.exceptions.InvalidFileException as ife:
        raise ValueError(
            f'Invalid file format or corrupted file: {input_file} - {str(ife)}'
            )
    except zipfile.BadZipFile as bzb:
        raise ValueError(
            f'File is not a valid Excel file or is corrupted: {input_file} - {str(bzb)}'
            )
    except AttributeError as ae:
        raise ValueError(
            f'Error processing file: {str(ae)}, File: {input_file}. Check Pillow version for compatibility regarding "ANTIALIAS".'
            )
    except Exception as e:
        raise ValueError(
            f'Error processing file: {str(e)}, File: {input_file}, Make sure it is a valid Excel file with the correct format.'
            ) from e
    finally:
        if workbook is not None:
            workbook.close()
        cleanup_browser()


if __name__ == '__main__':
    app = ScraperApp()
    app.run()
