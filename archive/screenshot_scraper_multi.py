import pandas as pd
import os
import argparse
import healing_agent
from search_engines import Google, Bing, Yahoo, Aol, Duckduckgo, Startpage
import openpyxl
from playwright.sync_api import sync_playwright, TimeoutError
import re
import time
from datetime import datetime
import random
import zipfile
import logging
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

logging.basicConfig(level=logging.INFO, format=
    '%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)
thread_local = threading.local()


def get_browser():
    if not hasattr(thread_local, 'browser'):
        playwright = sync_playwright().start()
        browser = playwright.firefox.launch()
        context = browser.new_context(viewport={'width': 1280, 'height': 
            1280}, user_agent=
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36'
            , device_scale_factor=1, has_touch=False, is_mobile=False,
            java_script_enabled=True)
        thread_local.playwright = playwright
        thread_local.browser = browser
        thread_local.context = context
    return thread_local.context


def cleanup_thread_browser():
    if hasattr(thread_local, 'browser'):
        thread_local.browser.close()
        thread_local.playwright.stop()
        del thread_local.browser
        del thread_local.context
        del thread_local.playwright


@healing_agent
def get_official_url(company_name, city, search_engine_starters=[Google(), Bing(), Yahoo(), Aol(), Duckduckgo(), Startpage()]):
    """
    Get the official website URL for a company by searching for it online.
    Uses multiple search engines and filters out blacklisted domains.
    
    Args:
        company_name: Name of the company
        city: City where company is located
        
    Returns:
        str: Official website URL if found, "na" if blocked, None otherwise
    """
    blacklist = ['opten.hu', 'facebook.com', 'linkedin.com',
        'instagram.com', 'ceginformacio.hu', 'ceginfo.hu', 'twitter.com',
        'youtube.com', 'pinterest.com', 'glassdoor.com', 'indeed.com',
        'bloomberg.com', 'crunchbase.com', 'yelp.com', 'yellowpages.com',
        'bbb.org', 'manta.com', 'bizapedia.com', 'google.hu', 'google.com',
        'cylex.hu', 'nemzeticegtar.hu', 'onlinecegnyilvantarto.hu',
        'wikipedia.org', 'companywall.hu', 'emis.com', 'economx.hu',
        'portfolio.hu', 'mfor.hu', 'kozlonyok.hu', 'trademagazin.hu',
        'magyarcegnyilvantarto.hu', 'topcegkereso.hu', 'xir.hu',
        'cegcontrol.hu', 'cegtalalo.hu', 'k-monitor.hu', 'dnb.com']
    blocklist = ['google.hu', 'google.com']
    query = f'{company_name} hivatalos weblap'
    logger.info(f'Searching for {query}')
    for engine in search_engine_starters:
        try:
            engine.ignore_duplicate_urls = True
            engine.ignore_duplicate_domains = True
            results = engine.search(query, pages=1)
            urls = results.links()
            if not urls:
                return 'na'
            filtered_urls = [url for url in urls if not any(blocked in url.
                lower() for blocked in blocklist)]
            if not filtered_urls:
                return 'na'
            urls = filtered_urls
            first_url = urls[0]
            if any(blocked in first_url.lower() for blocked in blacklist):
                logger.info(
                    f'First result for {company_name} was blocked: {first_url}'
                    )
                return 'na'
            parts = first_url.split('/')
            if len(parts) > 3 and 'hu' in parts[3]:
                main_url = '/'.join(parts[:4])
            else:
                main_url = '/'.join(parts[:3])
            logger.info(f'Found URL for {company_name}: {main_url}')
            return main_url
        except Exception as e:
            logger.error(f'Error with {engine.__class__.__name__}: {str(e)}')
            continue
    return None


@healing_agent
def get_screenshot_url(official_url, row_number):
    """
    Navigate to URL, handle cookie popups, take screenshot and save it using Playwright.
    Includes anti-blocking measures and detection.
    
    Args:
        official_url: Website URL to screenshot
        row_number: Excel row number for filename
        
    Returns:
        str: Screenshot filename if successful, None otherwise
    """
  
    start_time = time.time()
    page = None
    if not official_url:
        return None
    screenshots_dir = 'screenshots'
    if not os.path.exists(screenshots_dir):
        os.makedirs(screenshots_dir)
    domain = official_url.split('//')[-1].split('/')[0]
    clean_domain = re.sub('[^\\w\\-_]', '', domain)
    filename = f'{row_number}_{clean_domain}.jpg'
    filepath = os.path.join(screenshots_dir, filename)
    try:
        logger.info(
            f'Starting browser setup at {time.time() - start_time:.2f}s')
        context = get_browser()
        page = context.new_page()
        logger.info(
            f'Browser setup completed at {time.time() - start_time:.2f}s')

        def handle_route(route):
            headers = route.request.headers
            headers['Accept'] = (
                'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'
                )
            headers['Accept-Language'] = 'hu-HU,hu;q=0.9,en;q=0.8'
            route.continue_(headers=headers)
        page.route('**/*', handle_route)
        logger.info(f'Starting page load at {time.time() - start_time:.2f}s')
        try:
            page.goto(official_url, wait_until='networkidle', timeout=10000)
        except TimeoutError:
            logger.warning(
                f'Networkidle timeout reached for {official_url}, taking screenshot of current state'
                )
        logger.info(f'Page loaded at {time.time() - start_time:.2f}s')
        cookie_selectors = ["button:has-text('Accept')",
            "button:has-text('Elfogad')", "button:has-text('Elfogadom')",
            "button:has-text('Rendben')", "button:has-text('Megértettem')",
            "button:has-text('Hozzájárulok')", "button:has-text('Bezár')",
            "button:has-text('Accept Cookies')",
            "button:has-text('I Accept')", "button:has-text('Allow All')",
            "button:has-text('Allow Cookies')", "button:has-text('süti')",
            "button:has-text('Süti')", "button:has-text('OK')",
            "button:has-text('ÖSSZES')", "button:has-text('Összes')",
            "button:has-text('Összes süti elfogadása')",
            "button:has-text('OK')", "button:has-text(' OK ')",
            "button:has-text('Elfogadása')", "button:has-text('Elfogadom')",
            "button:has-text('Rendben')", "button:has-text('Megértettem')",
            "button:has-text('Hozzájárulok')", "button:has-text('Bezár')",
            "button:has-text('Mindent elfogadok')",
            "button:has-text('Alle Akzeptieren')",
            "button:has-text('ELFOGADOM A JAVASOLT BEÁLLÍTÁSOKAT')",
            "button:has-text('AGREE AND CLOSE')",
            "button:has-text('Beállítások mentése')",
            "button:has-text('Az összes süti elfogadása')",
            "button:has-text('Javasolt beállítások elfogadása')",
            "button:has-text('Confirm')", '[id*=cookie] button',
            '[class*=cookie] button', '[id*=cookie-accept]',
            '[class*=cookie-accept]', '[id*=cookie-consent] button',
            '[class*=cookie-consent] button', '[id*=cookieConsent] button',
            '[class*=cookieConsent] button', '[id*=cookie-banner] button',
            '[class*=cookie-banner] button', '[id*=cookie-policy] button',
            '[class*=cookie-policy] button', '[id*=gdpr] button',
            '[class*=gdpr] button', '[id*=privacy] button',
            '[class*=privacy] button', "[aria-label*='cookie'] button",
            '[data-cookiebanner] button']
        logger.info(
            f'Starting cookie handling at {time.time() - start_time:.2f}s')
        for selector in cookie_selectors:
            try:
                if page.is_visible(selector):
                    page.click(selector, timeout=100)
                    break
            except Exception as ex:
                logger.debug(
                    f'Cookie selector "{selector}" not visible: {str(ex)}')
                continue
        logger.info(
            f'Cookie handling completed at {time.time() - start_time:.2f}s')
        page.wait_for_timeout(1000 + int(1000 * random.random()))
        logger.info(f'Taking screenshot at {time.time() - start_time:.2f}s')
        page.screenshot(path=filepath)
        logger.info(
            f'Screenshot saved to {filepath} at {time.time() - start_time:.2f}s'
            )
        return filename
    except Exception as e:
        logger.error(f'Screenshot error for {official_url}: {str(e)}')
        return None
    finally:
        if page is not None:
            page.close()


def process_row(row_data):
    (current_row, company_name, city, official_url, screenshot_url,
        update_urls, update_screenshots) = row_data
    updates = {}
    if update_urls and not official_url:
        start_time = time.time()
        logger.info(f'Getting official URL for {company_name}...')
        new_url = get_official_url(company_name, city)
        elapsed = time.time() - start_time
        logger.info(f'Getting official URL took {elapsed:.2f} seconds')
        if new_url:
            updates['official_url'] = new_url
            official_url = new_url
    if update_screenshots and not screenshot_url and official_url:
        start_time = time.time()
        logger.info(f'Getting screenshot for {official_url}...')
        new_screenshot = get_screenshot_url(official_url, current_row)
        elapsed = time.time() - start_time
        logger.info(f'Getting screenshot took {elapsed:.2f} seconds')
        if new_screenshot:
            updates['screenshot_url'] = new_screenshot
    return current_row, updates


@healing_agent
def process_companies_file(input_file, update_urls=True, update_screenshots
    =True):
    """
    Process Excel file containing company information and add official_url
    and screenshot_url columns where missing. Uses openpyxl for better performance.
    
    Args:
        input_file: Path to Excel file
        update_urls: Whether to update missing official URLs
        update_screenshots: Whether to update missing screenshot URLs
    """
    if not os.path.exists(input_file):
        raise FileNotFoundError(f'Input file not found: {input_file}')
    try:
        workbook = openpyxl.load_workbook(input_file)
        workbook.close()
    except (openpyxl.utils.exceptions.InvalidFileException, zipfile.BadZipFile
        ) as e:
        logger.error(f'Invalid Excel file: {input_file}')
        logger.error(f'Error: {str(e)}')
        return None
    workbook = None
    try:
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        official_url_idx = headers.index('official_url'
            ) if 'official_url' in headers else None
        screenshot_url_idx = headers.index('screenshot_url'
            ) if 'screenshot_url' in headers else None
        company_name_idx = headers.index('cégnév')
        city_idx = headers.index('város')
        if official_url_idx is None:
            official_url_idx = len(headers)
            headers.append('official_url')
            sheet.cell(row=1, column=official_url_idx + 1, value='official_url'
                )
        if screenshot_url_idx is None:
            screenshot_url_idx = len(headers)
            headers.append('screenshot_url')
            sheet.cell(row=1, column=screenshot_url_idx + 1, value=
                'screenshot_url')
        total_rows = sheet.max_row
        logger.info(f'Total rows: {total_rows}')
        thread1_rows = range(2, total_rows + 1, 2)
        thread2_rows = range(3, total_rows + 1, 2)
        with ThreadPoolExecutor(max_workers=2
            ) as url_executor, ThreadPoolExecutor(max_workers=2
            ) as screenshot_executor:
            for current_row in [thread1_rows, thread2_rows]:
                updates_made = False
                for row_idx in current_row:
                    company_name = sheet.cell(row=row_idx, column=
                        company_name_idx + 1).value
                    city = sheet.cell(row=row_idx, column=city_idx + 1).value
                    official_url = sheet.cell(row=row_idx, column=
                        official_url_idx + 1).value
                    screenshot_url = sheet.cell(row=row_idx, column=
                        screenshot_url_idx + 1).value
                    if official_url == 'na' and not screenshot_url:
                        sheet.cell(row=row_idx, column=screenshot_url_idx +
                            1, value='na')
                        updates_made = True
                        continue
                    if update_urls and not official_url:
                        google_future_engines = [Google(), Yahoo()]
                        bing_future_engines = [Bing(), Aol()]
                        google_future = url_executor.submit(get_official_url,
                            company_name, city, google_future_engines)
                        bing_future = url_executor.submit(get_official_url,
                            company_name, city, bing_future_engines)
                        for future in as_completed([google_future, bing_future]
                            ):
                            new_url = future.result()
                            if new_url:
                                official_url = new_url
                                sheet.cell(row=row_idx, column=
                                    official_url_idx + 1, value=new_url)
                                updates_made = True
                                break
                    if (update_screenshots and not screenshot_url and
                        official_url and official_url != 'na'):
                        screenshot_future1 = screenshot_executor.submit(
                            get_screenshot_url, official_url, row_idx)
                        screenshot_future2 = screenshot_executor.submit(
                            get_screenshot_url, official_url, row_idx)
                        screenshots = []
                        for future in as_completed([screenshot_future1,
                            screenshot_future2]):
                            new_screenshot = future.result()
                            if new_screenshot:
                                screenshots.append(new_screenshot)
                        if screenshots:
                            sheet.cell(row=row_idx, column=
                                screenshot_url_idx + 1, value=
                                ','.join(screenshots))
                            updates_made = True
                if updates_made:
                    workbook.save(input_file)
        return pd.read_excel(input_file)
    except FileNotFoundError as fnf_error:
        raise FileNotFoundError(f'Input file issue: {str(fnf_error)}')
    except Exception as e:
        logger.error(f'Error processing file: {str(e)}')
        return None
    finally:
        if workbook is not None:
            workbook.close()
        cleanup_thread_browser()


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', default=
        'ceglista_2024_09_16 - Copy.xlsx', help='Input Excel file path')
    parser.add_argument('--urls-only', action='store_true', help=
        'Only update missing official URLs')
    parser.add_argument('--screenshots-only', action='store_true', help=
        'Only update missing screenshot URLs')
    args = parser.parse_args()
    update_urls = not args.screenshots_only
    update_screenshots = not args.urls_only
    process_companies_file(args.input, update_urls, update_screenshots)
